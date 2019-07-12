#病人类
from datetime import datetime ,date ,timedelta
# from xlutils.copy import copy
import openpyxl
import xlrd 
import xlwt
# workingmatrix=[[ 7, 7, 7, 7, 7, 7, 7],  # 第二问矩阵
#                [13,12,12,12,12,13,12],
#                [11,10,10,10,10,11,10],
#                [12,11,10, 9, 8, 6, 5],
#                [ 5, 4, 8, 7, 6, 5, 4]]

workingmatrix=[[ 7, 7, 7, 7, 9, 8, 7],  # 第四问矩阵
               [13,12,12,15,14,13,12],
               [11,10,10,13,12,11,10],
               [12,11,10, 9, 8, 6, 5],
               [ 5, 4, 8, 7, 6, 5, 4]]
names = {
  "外伤" : 0,
  "视网膜疾病" : 1,
  "青光眼" : 2,
  "白内障(双眼)" : 3, 
  "白内障" : 4
}

class Sickman:
   def __init__(self,num,name,starttime):
      self.num = num   #对应number序号好填表
      self.name = name 
      self.starttime = starttime 
      week = starttime.weekday()
      nameindex = names[name]
      self.lasttime = workingmatrix[nameindex][week]
      self.priority = 0 
      self.endtime = None
   
   def getNum(self):
      return self.num
   def getName(self):
      return self.name
   def getStarttime(self):
      return self.starttime
   def update_priority(self,currenttime):
      #公式
      week = self.starttime.weekday()
      nameindex = names[self.name]
      s_next = workingmatrix[nameindex][(week+1)%7]
      s = self.lasttime
      w = (currenttime - self.starttime + timedelta(days=1)).days
      self.priority =  s_next/s*(1+w/s)
      
   def __lt__(self, other):
      return self.priority> other.priority

   def Update(self,currenttime,workbook,path):
      # 当人进病房里的时候，更新endtime
      endtime = currenttime +timedelta(days=self.lasttime) # 注意是date运算
      #应该在主程序中一开始就写好,减少IO
      #workbook = xlrd.open_workbook(path)
      new_sheet = workbook[workbook.sheetnames[0]]
      # dateFormat = xlwt.XFStyle()
      # dateFormat.num_format_str = 'yyyy/mm/dd'
      # new_sheet.write(self.num,7,currenttime,dateFormat) #入院时间
      # new_sheet.write(self.num,8,endtime,dateFormat)     #出院时间   #按序号写行,写到第七列
      new_sheet.cell(self.num+1,8).value = currenttime
      new_sheet.cell(self.num+1,9).value = endtime
      workbook.save(path)
      print("写入成功"+str(endtime))
   